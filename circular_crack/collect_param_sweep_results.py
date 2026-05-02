#!/usr/bin/env python3
"""
Collect post-processed dynamic parameter-sweep results into comparison workbooks.

For each requested velocity and sweep group, this script reads the case folders
listed in param_sweep_v*_summary.csv and creates one xlsx workbook with four
sheets:

1. local_stress_norm_ave
2. local_stress_norm_max_min
3. local_stress_ring_5hL
4. dsif_normalized
"""

import argparse
import csv
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - exercised only without dependency
    raise SystemExit(
        "openpyxl is required to write xlsx files. "
        "Install it with: python3 -m pip install openpyxl"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_RESULTS_DIR = SCRIPT_DIR / "results" / "param_sweep_dynamic"
DEFAULT_OUTPUT_DIRNAME = "collected_postprocess"

GROUPS = ("rGL", "aL", "lL", "HL")
PROFILE_FILE = "local_stress_normalized_profile.csv"
RING5_FILE = "local_stress_ring_5hL.csv"
DSIF_FILE = "dsif_normalized.csv"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="EAF4EA")


def format_velocity(value):
    value_float = float(value)
    if value_float.is_integer():
        return str(int(value_float))
    return f"{value_float:g}".replace(".", "p")


def read_csv_dicts(path):
    with Path(path).open(newline="") as f:
        return list(csv.DictReader(f))


def to_number(value):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return int(number)
    return number


def numeric_key(value):
    number = to_number(value)
    if isinstance(number, (int, float)):
        return float(number)
    return None


def read_summary(results_dir, velocity):
    v_label = format_velocity(velocity)
    summary_path = results_dir / f"param_sweep_v{v_label}_summary.csv"
    if not summary_path.exists():
        raise FileNotFoundError(f"Missing summary CSV: {summary_path}")
    return read_csv_dicts(summary_path), summary_path


def case_label(row):
    group = row.get("group", "")
    if group in row and row[group] != "":
        return f"{group}={row[group]}"
    return row.get("label") or f"idx={row.get('idx', '')}"


def case_folder(row, results_dir):
    folder_text = row.get("folder", "")
    if not folder_text:
        return None
    folder = Path(folder_text)
    if not folder.is_absolute():
        folder = results_dir / folder
    return folder


def load_case_outputs(row, results_dir):
    folder = case_folder(row, results_dir)
    if folder is None:
        return {
            "row": row,
            "label": case_label(row),
            "folder": None,
            "profile": [],
            "ring5": [],
            "dsif": [],
            "missing": ["folder"],
        }

    postprocess_dir = folder / "postprocess"
    paths = {
        "profile": postprocess_dir / PROFILE_FILE,
        "ring5": postprocess_dir / RING5_FILE,
        "dsif": postprocess_dir / DSIF_FILE,
    }

    missing = []
    outputs = {}
    for name, path in paths.items():
        if path.exists():
            outputs[name] = read_csv_dicts(path)
        else:
            outputs[name] = []
            missing.append(str(path))

    return {
        "row": row,
        "label": case_label(row),
        "folder": folder,
        "profile": outputs["profile"],
        "ring5": outputs["ring5"],
        "dsif": outputs["dsif"],
        "missing": missing,
    }


def group_cases(summary_rows, groups):
    grouped = defaultdict(list)
    for row in summary_rows:
        group = row.get("group")
        if group in groups:
            grouped[group].append(row)
    return grouped


def collect_profile_distances(cases):
    distances = {}
    for case in cases:
        for row in case["profile"]:
            distance = numeric_key(row.get("distance_hL"))
            if distance is None:
                continue
            distance_mm = to_number(row.get("distance_mm"))
            distances.setdefault(distance, distance_mm)
    return sorted(distances.items(), key=lambda item: item[0])


def profile_value_map(case, column):
    values = {}
    for row in case["profile"]:
        distance = numeric_key(row.get("distance_hL"))
        if distance is not None:
            values[distance] = to_number(row.get(column))
    return values


def write_profile_ave_sheet(wb, cases):
    ws = wb.create_sheet("local_stress_norm_ave")
    distances = collect_profile_distances(cases)
    value_maps = [profile_value_map(case, "norm_ave") for case in cases]

    headers = ["distance_hL", "distance_mm"] + [case["label"] for case in cases]
    ws.append(headers)
    style_header_row(ws, 1)

    for distance, distance_mm in distances:
        row = [format_distance(distance), distance_mm]
        row.extend(value_map.get(distance) for value_map in value_maps)
        ws.append(row)

    freeze_and_autofit(ws, "C2")


def write_profile_max_min_sheet(wb, cases):
    ws = wb.create_sheet("local_stress_norm_max_min")
    distances = collect_profile_distances(cases)
    max_maps = [profile_value_map(case, "norm_max") for case in cases]
    min_maps = [profile_value_map(case, "norm_min") for case in cases]

    headers = ["distance_hL", "metric"] + [case["label"] for case in cases]
    ws.append(headers)
    style_header_row(ws, 1)

    for distance, _distance_mm in distances:
        max_row = [format_distance(distance), "norm_max"]
        max_row.extend(value_map.get(distance) for value_map in max_maps)
        ws.append(max_row)

        min_row = [format_distance(distance), "norm_min"]
        min_row.extend(value_map.get(distance) for value_map in min_maps)
        ws.append(min_row)

        ws.append([])

    freeze_and_autofit(ws, "C2")


def write_pair_sheet(wb, title, cases, source_key):
    ws = wb.create_sheet(title)

    for case_index, case in enumerate(cases):
        start_col = 1 + 2 * case_index
        end_col = start_col + 1
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=end_col,
        )
        cell = ws.cell(row=1, column=start_col, value=case["label"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL

        ws.cell(row=2, column=start_col, value="theta_deg")
        ws.cell(row=2, column=end_col, value="normalized")
        ws.cell(row=2, column=start_col).fill = SUBHEADER_FILL
        ws.cell(row=2, column=end_col).fill = SUBHEADER_FILL
        ws.cell(row=2, column=start_col).font = Font(bold=True)
        ws.cell(row=2, column=end_col).font = Font(bold=True)

        for row_index, row in enumerate(case[source_key], start=3):
            ws.cell(row=row_index, column=start_col, value=to_number(row.get("theta_deg")))
            ws.cell(row=row_index, column=end_col, value=to_number(row.get("normalized")))

    freeze_and_autofit(ws, "A3")


def format_distance(distance):
    if float(distance).is_integer():
        return int(distance)
    return distance


def style_header_row(ws, row_index):
    for cell in ws[row_index]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def freeze_and_autofit(ws, freeze_pane):
    ws.freeze_panes = freeze_pane
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 28)


def write_group_workbook(output_path, cases):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_profile_ave_sheet(wb, cases)
    write_profile_max_min_sheet(wb, cases)
    write_pair_sheet(wb, "local_stress_ring_5hL", cases, "ring5")
    write_pair_sheet(wb, "dsif_normalized", cases, "dsif")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def collect_velocity(results_dir, output_dir, velocity, groups, strict=False):
    v_label = format_velocity(velocity)
    summary_rows, summary_path = read_summary(results_dir, velocity)
    grouped = group_cases(summary_rows, groups)

    written = []
    warnings = []
    for group in groups:
        rows = grouped.get(group, [])
        if not rows:
            message = f"v{v_label} group {group}: no rows in {summary_path}"
            if strict:
                raise RuntimeError(message)
            warnings.append(message)
            continue

        cases = [load_case_outputs(row, results_dir) for row in rows]
        for case in cases:
            for missing in case["missing"]:
                warnings.append(f"v{v_label} {group} {case['label']}: missing {missing}")

        output_path = output_dir / f"param_sweep_v{v_label}_{group}_postprocess.xlsx"
        write_group_workbook(output_path, cases)
        written.append(output_path)

    return written, warnings


def parse_args():
    parser = argparse.ArgumentParser(
        description="Collect dynamic parameter-sweep postprocess CSV files into xlsx workbooks."
    )
    parser.add_argument(
        "--results-dir",
        type=Path,
        default=DEFAULT_RESULTS_DIR,
        help="Directory containing param_sweep_v*_summary.csv and v*/case folders.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help=(
            "Directory for collected workbooks. "
            f"Default: RESULTS_DIR/{DEFAULT_OUTPUT_DIRNAME}"
        ),
    )
    parser.add_argument(
        "--velocities",
        nargs="+",
        default=["500", "1000"],
        help="Velocities to collect, e.g. --velocities 500 1000.",
    )
    parser.add_argument(
        "--groups",
        nargs="+",
        default=list(GROUPS),
        choices=list(GROUPS),
        help="Sweep groups to collect.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail if any requested group is missing from a summary CSV.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    results_dir = args.results_dir.resolve()
    output_dir = (
        args.output_dir.resolve()
        if args.output_dir is not None
        else results_dir / DEFAULT_OUTPUT_DIRNAME
    )

    all_written = []
    all_warnings = []
    for velocity in args.velocities:
        written, warnings = collect_velocity(
            results_dir=results_dir,
            output_dir=output_dir,
            velocity=velocity,
            groups=args.groups,
            strict=args.strict,
        )
        all_written.extend(written)
        all_warnings.extend(warnings)

    print(f"Collected workbooks written to: {output_dir}")
    for path in all_written:
        print(f"  - {path}")

    if all_warnings:
        print("\nWarnings:")
        for warning in all_warnings:
            print(f"  - {warning}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
